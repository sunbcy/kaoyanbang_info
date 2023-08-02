import pandas as pd
import numpy as np
from openpyxl.utils import get_column_letter

xls_header_map = {
    'school_id': {'index': 1, 'zh': '学校ID'},
    'school_code': {'index': 2, 'zh': '学校代码'},
    'school_name': {'index': 3, 'zh': '学校名称'},
    'is_211': {'index': 4, 'zh': '211'},
    'is_985': {'index': 5, 'zh': '985'},
    'is_score': {'index': 6, 'zh': '自主划线'},
    'is_first_class': {'index': 7, 'zh': '双一流'},
    'is_other': {'index': 8, 'zh': '普通高校'},
    'follow_num': {'index': 9, 'zh': '关注人数'},
    'exam_num': {'index': 10, 'zh': '今年备考人数'},
    'week_exam_num': {'index': 11, 'zh': '近7日新增'},
    'departments': {'index': 12, 'zh': '学院'},
    'province_id': {'index': 13, 'zh': '省份ID'},
    'province_name': {'index': 14, 'zh': '所在地区'},
    'school_type': {'index': 15, 'zh': '学校类型'},
    'is_edu': {'index': 16, 'zh': '教育部直属'},
    'is_local': {'index': 17, 'zh': '地方所属'},
    'is_center': {'index': 18, 'zh': '中央部委直属'},
    'bbs_id': {'index': 19, 'zh': '论坛ID'},
    'student_recruitment': {'index': 20, 'zh': '招生通告'},
    'badge': {'index': 21, 'zh': '学校校徽'},
}


def write_data_into_excel(xlspath, data_json_list, sheet_name='Sheet1'):
    writer = pd.ExcelWriter(xlspath, engine='openpyxl')
    sheetnames_origin = data_json_list[0].keys()  # 获取所有sheet(data_json_list中单元数据的键key)的名称
    sheetnames_en = []

    flag = 1
    while flag <= len(xls_header_map):  # 调控xlsx的顺序
        for item in xls_header_map:  # sheetnames_en
            if xls_header_map.get(item)['index'] == flag:
                sheetnames_en.append(item)
        flag += 1

    sheetnames_zh = [xls_header_map.get(sheet)['zh'] for sheet in sheetnames_en if xls_header_map.get(sheet)['zh']]
    print(sheetnames_zh)
    data = {}
    for sheet in sheetnames_zh:
        data[sheet] = []
    for sheet in sheetnames_en:
        for i in data_json_list:
            data[xls_header_map.get(sheet)['zh'] if xls_header_map.get(sheet)['zh'] else sheet].append(i[sheet])
    data = pd.DataFrame(data)
    data = data[sheetnames_zh]  # 将data按sheetnames_zh顺序排列
    data.to_excel(writer, sheet_name=sheet_name, index=False)

    # 计算每列表头的字符宽度
    column_widths = (data.columns.to_series().apply(lambda x: len(str(x).encode('gbk'))).values)
    # 计算每列的最大字符宽度
    max_widths = (data.astype(str).applymap(lambda x: len(str(x).encode('gbk'))).agg(max).values)
    # 取两者中每列的最大宽度
    widths = np.max([column_widths, max_widths], axis=0)
    # 指定sheet, 设置该sheet的每列列宽
    worksheet = writer.sheets[sheet_name]
    for i, width in enumerate(widths, 1):
        # openpyxl引擎设置字符宽度时会缩水0.5个左右字符串,所以干脆+2使左右都空出一个字宽.
        worksheet.column_dimensions[get_column_letter(i)].width = width + 2
    # 保存writer中的数据至excel
    writer._save()  # 等于writer.close()
